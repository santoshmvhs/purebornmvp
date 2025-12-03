#!/usr/bin/env python3
"""
Script to import products from Excel file directly.
Usage: python scripts/import_products.py <excel_file_path>
"""
import sys
import os
import asyncio
from pathlib import Path

# Add parent directory to path to import app modules
sys.path.insert(0, str(Path(__file__).parent.parent))

# Set environment variable to skip validation during script execution
os.environ['ALEMBIC_CONTEXT'] = 'true'

from sqlalchemy.ext.asyncio import create_async_engine, AsyncSession
from sqlalchemy.orm import sessionmaker
from sqlalchemy import select, delete
from app.database import get_db
from app.models import Product, ProductCategory, ProductVariant
from app.config import settings
from decimal import Decimal
import re
import logging

try:
    import pandas as pd
    HAS_PANDAS = True
except ImportError:
    HAS_PANDAS = False
    try:
        from openpyxl import load_workbook
        HAS_OPENPYXL = True
    except ImportError:
        HAS_OPENPYXL = False
        print("Error: Need either pandas or openpyxl installed")
        print("Install with: pip install pandas openpyxl")
        sys.exit(1)

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


async def import_products_from_excel(file_path: str, db: AsyncSession):
    """Import products from Excel file."""
    
    if not os.path.exists(file_path):
        print(f"Error: File not found: {file_path}")
        return
    
    if not file_path.endswith(('.xlsx', '.xls')):
        print(f"Error: File must be an Excel file (.xlsx or .xls)")
        return
    
    try:
            # Read Excel file
            if HAS_PANDAS:
                df = pd.read_excel(file_path, header=0)
                if df.empty:
                    print("Error: Excel file is empty")
                    return
                # Normalize column names
                df.columns = df.columns.astype(str).str.strip().str.lower().str.replace(' ', '_').str.replace('-', '_').str.replace('.', '')
                df = df.dropna(how='all')
            else:
                # Use openpyxl
                wb = load_workbook(file_path, data_only=True)
                ws = wb.active
                
                # Read headers from first row
                headers = [str(cell.value).strip().lower().replace(' ', '_').replace('-', '_').replace('.', '') 
                          for cell in ws[1]]
                
                # Read data rows
                rows_data = []
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if any(cell is not None for cell in row):
                        rows_data.append(row)
                
                if not rows_data:
                    print("Error: Excel file is empty")
                    return
                
                # Convert to DataFrame-like structure
                import collections
                df = collections.namedtuple('DataFrame', ['columns', 'iterrows'])
                df.columns = headers
                df.iterrows = lambda: enumerate(rows_data)
                
                # Create a simple wrapper for pandas-like access
                class SimpleDF:
                    def __init__(self, headers, rows):
                        self.headers = headers
                        self.rows = rows
                        self.columns = headers
                    
                    def __len__(self):
                        return len(self.rows)
                    
                    def iterrows(self):
                        for idx, row in enumerate(self.rows):
                            yield idx, RowWrapper(row, self.headers)
                
                class RowWrapper:
                    def __init__(self, row, headers):
                        self._row = row
                        self._headers = headers
                    
                    def __getitem__(self, key):
                        if isinstance(key, int):
                            return self._row[key] if key < len(self._row) else None
                        # Find column index
                        try:
                            key_normalized = key.lower().strip().replace(' ', '_').replace('-', '_').replace('.', '')
                            idx = self._headers.index(key_normalized)
                            return self._row[idx] if idx < len(self._row) else None
                        except ValueError:
                            return None
                
                df = SimpleDF(headers, rows_data)
            
            print(f"Importing products from: {file_path}")
            print(f"Found {len(df)} rows")
            print(f"Columns: {list(df.columns)}")
            
            # Map column names
            column_mapping = {
                'category': ['category'],
                'sub_category': ['sub_category', 'subcategory'],
                'item_name': ['item_name', 'itemname', 'product_name', 'product'],
                'variant_name': ['variant_name', 'variantname', 'variant'],
                'product_code': ['product_code', 'productcode', 'product_code_alias'],
                'sku_code': ['sku_code', 'skucode', 'sku'],
                'metric': ['metric', 'unit'],
                'selling_price': ['selling_price', 'sellingprice', 'price'],
                'channel': ['channel'],
                'tax_class': ['tax_class', 'taxclass'],
            }
            
            # Find actual column names
            actual_columns = {}
            for key, variations in column_mapping.items():
                for col in df.columns:
                    col_normalized = col.lower().strip()
                    if col_normalized in variations:
                        actual_columns[key] = col
                        break
            
            # Validate required columns
            required = ['item_name', 'variant_name']
            missing = [col for col in required if col not in actual_columns]
            if missing:
                print(f"Error: Missing required columns: {missing}")
                print(f"Found columns: {list(df.columns)}")
                return
            
            # Helper functions
            def normalize_unit(unit_str):
                if not unit_str or (HAS_PANDAS and pd.isna(unit_str)) or str(unit_str).strip().lower() in ['nan', 'none', '']:
                    return 'Unit'
                unit = str(unit_str).strip().lower()
                unit_map = {
                    'ml': 'L', 'litre': 'L', 'liter': 'L', 'l': 'L',
                    'kg': 'kg', 'gram': 'kg', 'g': 'kg',
                    'unit': 'Unit', 'btl': 'Unit', 'piece': 'Unit', 'pcs': 'Unit',
                }
                for key, value in unit_map.items():
                    if key in unit:
                        return value
                return 'Unit'
            
            def extract_multiplier(variant_name, base_unit):
                if not variant_name:
                    return Decimal(1)
                variant_lower = str(variant_name).lower()
                numbers = re.findall(r'\d+\.?\d*', variant_lower)
                if numbers:
                    num = Decimal(numbers[0])
                    if 'ml' in variant_lower and base_unit == 'L':
                        return num / Decimal(1000)
                    elif ('gram' in variant_lower or 'g' in variant_lower) and base_unit == 'kg':
                        return num / Decimal(1000)
                return Decimal(1)
            
            # Load existing data
            from sqlalchemy import select, func
            
            result = await db.execute(select(ProductCategory))
            existing_categories = {cat.name.lower(): cat for cat in result.scalars().all()}
            
            result = await db.execute(select(Product))
            existing_products = {p.product_code: p for p in result.scalars().all() if p.product_code}
            
            result = await db.execute(select(ProductVariant))
            existing_variants = {v.sku: v for v in result.scalars().all() if v.sku}
            
            created_products = []
            created_variants = []
            updated_products = []
            updated_variants = []
            skipped = []
            errors = []
            
            # Process each row
            for idx, row in df.iterrows():
                try:
                    # Get category
                    category_name = None
                    if 'category' in actual_columns:
                        category_name = str(row[actual_columns['category']]).strip()
                        if category_name and category_name.lower() != 'nan':
                            category_name = category_name.strip()
                    
                    # Get or create category
                    category_id = None
                    if category_name:
                        category_key = category_name.lower()
                        if category_key in existing_categories:
                            category_id = existing_categories[category_key].id
                        else:
                            new_category = ProductCategory(name=category_name)
                            db.add(new_category)
                            await db.flush()
                            await db.refresh(new_category)
                            existing_categories[category_key] = new_category
                            category_id = new_category.id
                            print(f"  Created category: {category_name}")
                    
                    # Get product details
                    item_name = str(row[actual_columns['item_name']]).strip()
                    if not item_name or item_name.lower() == 'nan':
                        skipped.append({'row': idx + 1, 'reason': 'Missing item name'})
                        continue
                    
                    product_code = None
                    if 'product_code' in actual_columns:
                        code_val = row[actual_columns['product_code']]
                        if code_val is not None and (not HAS_PANDAS or not pd.isna(code_val)):
                            product_code = str(code_val).strip()
                            if product_code.lower() == 'nan' or product_code.lower() == 'none':
                                product_code = None
                    
                    # Get base unit
                    base_unit = 'Unit'
                    if 'metric' in actual_columns:
                        metric_val = row[actual_columns['metric']]
                        if metric_val is not None and (not HAS_PANDAS or not pd.isna(metric_val)):
                            base_unit = normalize_unit(str(metric_val))
                    
                    # Default HSN code
                    hsn_code = '00000000'
                    
                    # Get or create product
                    product = None
                    if product_code and product_code in existing_products:
                        product = existing_products[product_code]
                        if product.name != item_name:
                            product.name = item_name
                            updated_products.append(product_code)
                            print(f"  Updated product: {item_name}")
                    else:
                        product = Product(
                            name=item_name,
                            product_code=product_code,
                            category_id=category_id,
                            base_unit=base_unit,
                            hsn_code=hsn_code,
                            is_active=True
                        )
                        db.add(product)
                        await db.flush()
                        await db.refresh(product)
                        if product_code:
                            existing_products[product_code] = product
                        created_products.append(item_name)
                        print(f"  Created product: {item_name}")
                    
                    # Get variant details
                    variant_name = str(row[actual_columns['variant_name']]).strip()
                    if not variant_name or variant_name.lower() == 'nan':
                        variant_name = item_name
                    
                    sku = None
                    if 'sku_code' in actual_columns:
                        sku_val = row[actual_columns['sku_code']]
                        if sku_val is not None and (not HAS_PANDAS or not pd.isna(sku_val)):
                            sku = str(sku_val).strip()
                            if sku.lower() == 'nan' or sku.lower() == 'none':
                                sku = None
                    
                    selling_price = None
                    if 'selling_price' in actual_columns:
                        price_val = row[actual_columns['selling_price']]
                        if price_val is not None and (not HAS_PANDAS or not pd.isna(price_val)):
                            try:
                                selling_price = Decimal(str(price_val))
                            except:
                                pass
                    
                    channel = None
                    if 'channel' in actual_columns:
                        channel_val = row[actual_columns['channel']]
                        if channel_val is not None and (not HAS_PANDAS or not pd.isna(channel_val)):
                            channel = str(channel_val).strip()
                    
                    # Calculate multiplier
                    multiplier = extract_multiplier(variant_name, base_unit)
                    
                    # Get or create variant
                    variant = None
                    if sku and sku in existing_variants:
                        variant = existing_variants[sku]
                        if variant.product_id != product.id:
                            variant.product_id = product.id
                        if variant.variant_name != variant_name:
                            variant.variant_name = variant_name
                        if selling_price and variant.selling_price != selling_price:
                            variant.selling_price = selling_price
                        if channel and variant.channel != channel:
                            variant.channel = channel
                        updated_variants.append(sku)
                        print(f"    Updated variant: {variant_name}")
                    else:
                        variant = ProductVariant(
                            product_id=product.id,
                            variant_name=variant_name,
                            multiplier=float(multiplier),
                            sku=sku,
                            selling_price=float(selling_price) if selling_price else None,
                            channel=channel,
                            is_active=True
                        )
                        db.add(variant)
                        await db.flush()
                        await db.refresh(variant)
                        if sku:
                            existing_variants[sku] = variant
                        created_variants.append(variant_name)
                        print(f"    Created variant: {variant_name}")
                    
                except Exception as e:
                    error_msg = f"Row {idx + 1}: {str(e)}"
                    logger.error(error_msg, exc_info=True)
                    errors.append(error_msg)
                    print(f"  Error on row {idx + 1}: {e}")
            
            await db.commit()
            
            print("\n" + "=" * 60)
            print("Import Summary:")
            print("=" * 60)
            print(f"✓ Created: {len(created_products)} products, {len(created_variants)} variants")
            print(f"↻ Updated: {len(updated_products)} products, {len(updated_variants)} variants")
            print(f"⚠ Skipped: {len(skipped)} items")
            print(f"✗ Errors: {len(errors)}")
            
            if skipped:
                print("\nSkipped items:")
                for item in skipped[:10]:
                    print(f"  Row {item['row']}: {item['reason']}")
            
            if errors:
                print("\nErrors:")
                for error in errors[:10]:
                    print(f"  {error}")
            
            print("=" * 60)
            
    except Exception as e:
        logger.error(f"Error importing products: {e}", exc_info=True)
        print(f"\nError: {e}")
        raise


async def delete_all_products_and_variants(db: AsyncSession):
    """Delete all products and variants."""
    
    # Count before deletion
    result = await db.execute(select(Product))
    product_count = len(result.scalars().all())
    
    result = await db.execute(select(ProductVariant))
    variant_count = len(result.scalars().all())
    
    if product_count == 0 and variant_count == 0:
        print("No products or variants to delete.")
        return 0, 0
    
    print(f"\nDeleting {product_count} products and {variant_count} variants...")
    
    # Delete variants first (due to foreign key constraint)
    await db.execute(delete(ProductVariant))
    
    # Delete products
    await db.execute(delete(Product))
    
    await db.commit()
    
    print(f"✓ Deleted {product_count} products and {variant_count} variants\n")
    return product_count, variant_count


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser(description='Import products from Excel file')
    parser.add_argument('file_path', help='Path to Excel file (.xlsx or .xls)')
    parser.add_argument('--delete-first', action='store_true', help='Delete all existing products and variants before importing')
    args = parser.parse_args()
    
    file_path = args.file_path
    
    # Check if DATABASE_URL is set
    if not settings.DATABASE_URL or settings.DATABASE_URL == "postgresql://user:password@localhost:5432/posdb":
        print("Warning: DATABASE_URL not configured!")
        print("Please set DATABASE_URL in your .env file or as an environment variable")
        print("Example: export DATABASE_URL='postgresql+asyncpg://user:pass@localhost:5432/dbname'")
        sys.exit(1)
    
    async def run_import():
        # Create database connection
        db_url = settings.DATABASE_URL
        if db_url.startswith('postgresql://'):
            db_url = db_url.replace('postgresql://', 'postgresql+asyncpg://', 1)
        elif db_url.startswith('postgresql+psycopg'):
            db_url = db_url.replace('postgresql+psycopg', 'postgresql+asyncpg', 1)
        
        engine = create_async_engine(
            db_url,
            connect_args={"statement_cache_size": 0} if "asyncpg" in db_url else {}
        )
        async_session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)
        
        try:
            async with async_session() as db:
                if args.delete_first:
                    await delete_all_products_and_variants(db)
                # Import products from Excel
                await import_products_from_excel(file_path, db)
        finally:
            await engine.dispose()
    
    try:
        asyncio.run(run_import())
    except Exception as e:
        print(f"\nError: {e}")
        print("\nTroubleshooting:")
        print("1. Make sure DATABASE_URL is correct in .env file")
        print("2. Make sure the database is running and accessible")
        print("3. Make sure you have the required permissions")
        sys.exit(1)

